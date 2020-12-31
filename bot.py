import telebot
from telebot import types
import openpyxl
from openpyxl.styles import PatternFill
import requests
import webbrowser
from bs4 import BeautifulSoup as BS
import config

tovars = requests.get('http://procen.ru/ceny/')   # парсинг продуктов по ссылке
html = BS(tovars.content, 'html.parser')
tegs_a = html.findAll('a')[6:]
names_tovars = []
ssilki = []
prices = dict()
for i in range(len(tegs_a)):
    ssilka = tegs_a[i]['href'][5:]       # вытаскиваю ссылку на страницу товара
    ssilki.append(ssilka[1:])            # убираю слеш
    prices[tegs_a[i].text] = ssilka
    names_tovars.append(tegs_a[i].text)        # названия всех продуктов



bot = telebot.TeleBot(config.TOKEN)

food = {}  # словарь где будут название: {кол-во, ед.измерения, цена с сайта}
p = []
a = []


# список цен
m = ['249.0p/шт', '729.0p/кг', '105.0p/кг', '130.0p/кг', '109.0p/кг', '184.0p/кг', '65.0p/кг', '90.0p/кг', '62.0p/шт',
     '299.0p/шт', '630.0p/шт', '105.0p/кг', '2599.0p/шт', '41.5p/шт', '169.0p/шт', '30.0p/шт', '100.0p/шт', '89.0p/шт',
     '183.0p/шт', '55.0p/кг', '84.0p/кг', '78.0p/шт', '96.0p/шт', '150.0p/кг', '80.0p/шт', '20.0p/шт', '165.0p/кг',
     '25.0p/шт', '122.6p/шт', '177.6p/кг', '27.0p/шт', '110.0p/шт', '160.0p/кг', '26.0p/шт', '135.0p/кг', '170.0p/шт',
     '89.9p/шт', '100.0p/шт', '140.0p/шт', '30.0p/шт', '125.0p/кг', '309.0p/шт', '1729.0p/кг', '792.0p/кг', '120.0p/шт',
     '29.7p/шт', '250.0p/шт', '48.0p/шт', '45.0p/шт', '419.0p/кг', '60.0p/шт', '719.0p/шт', '129.0p/шт', '1069.9p/кг',
     '94.0p/шт', '36.0p/шт', '54.0p/шт', '70.0p/кг', '899.0p/шт', '130.0p/шт', '241.0p/шт', '62.5p/кг', '22.0p/кг',
     '48.0p/шт', '63.0p/шт', '240.0p/шт', '88.0p/шт', '82.0p/шт', '75.0p/шт', '50.0p/шт', '50.0p/шт', '299.0p/шт',
     '140.0p/шт', '140.0p/кг', '89.0p/шт', '470.0p/шт', '5.5p/шт', '73.0p/шт', '479.0p/шт', '150.0p/кг', '729.0p/кг',
     '89.0p/шт', '58.0p/шт', '329.0p/шт', '299.0p/шт', '105.0p/кг', '420.0p/шт', '54.0p/шт', '39.9p/шт', '65.0p/шт',
     '1399.0p/шт', '200.0p/шт', '46.0p/шт', '21.0p/шт', '20.0p/кг', '729.3p/кг', '56.0p/шт', '69.2p/шт', '89.0p/шт',
     '120.0p/шт', '119.0p/шт', '30.0p/кг', '111.0p/кг', '132.6p/шт', '103.0p/шт', '21.0p/шт', '21.0p/шт', '40.0p/шт',
     '24.5p/шт', '97.0p/шт', '45.0p/шт', '339.0p/шт', '50.0p/шт', '82.3p/шт', '268.0p/шт', '89.0p/шт', '37.0p/шт',
     '208.0p/шт', '52.5p/шт', '820.0p/кг', '56.0p/шт', '500.0p/шт', '88.0p/шт', '34.8p/шт', '78.0p/шт', '20.0p/шт',
     '129.0p/шт', '30.0p/кг', '32.0p/шт', '123.0p/кг', '62.0p/шт', '89.2p/кг', '114.9p/шт', '200.0p/шт', '129.0p/шт',
     '50.0p/шт', '28.0p/шт', '56.0p/шт', '170.0p/шт', '46.5p/шт', '129.0p/шт', '459.0p/шт', '50.0p/шт', '234.0p/шт',
     '15.9p/кг', '50.0p/шт', '86.0p/шт', '199.0p/шт', '250.0p/кг', '80.0p/шт', '68.0p/шт', '70.0p/шт', '50.0p/шт',
     '95.0p/кг', '560.0p/кг', '59.5p/шт']

# засовываю все в словарь food
for i in range(len(names_tovars)):
    food[names_tovars[i]] = [0, 'шт', m[i]]


@bot.message_handler(commands=['start'])  # приветствие
def welcome(message):
    sti = open('sticker.webp', 'rb')   # отправка стикера
    bot.send_sticker(message.chat.id, sti)
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)   # клавиатура с кнопками
    item1 = types.KeyboardButton('📋 Список продуктов')
    item2 = types.KeyboardButton('Сравнить цены')
    item3 = types.KeyboardButton('Необходимо купить')
    item4 = types.KeyboardButton('✍ EDIT')
    markup.add(item1, item2, item3, item4)
    bot.send_message(message.chat.id, 'Добро пожаловать, {0.first_name}!\nЯ - {1.first_name}, чем вам помочь?'.format(message.from_user, bot.get_me()), reply_markup=markup)

# реакции бота на ввод пользователя
@bot.message_handler(content_types=['text'])
def movie(message):
    if message.text == '📋 Список продуктов':
        print(food)
        if len(food) > 0:
            file_excel = openpyxl.load_workbook(filename='food1.xlsx', data_only=True)  # открываю файл ексель food1.xlsx
            sheet = file_excel['Sheet1']
            sheet['A1'], sheet['B1'], sheet['C1'], sheet['D1'], sheet['E1'] = 'Номер', 'Название продукта', 'Кол-во', 'Ед. Измерения', 'Цена'
            yellowFill = PatternFill(start_color='F9DA5C', end_color='F9DA5C', fill_type='solid')  # желтый цвет ячейки
            for i in range(len(names_tovars)):
                sheet['A' + str(i + 2)] = i + 1
                if int(food[names_tovars[i]][0]) == 0:
                    work_sheet = sheet['A' + str(i + 2)]
                    work_sheet.fill = yellowFill    # отмечаю ячейки желтым цветом если кол-во продукта равно 0
                sheet['B' + str(i + 2)] = names_tovars[i]
                sheet['C' + str(i + 2)] = int(food[names_tovars[i]][0])
                if len(food[names_tovars[i]]) == 2:
                    sheet['D' + str(i + 2)] = food[names_tovars[i]][1]
                elif len(food[names_tovars[i]]) == 3:
                    price = food[names_tovars[i]][2].split('/')
                    sheet['D' + str(i + 2)] = price[1]
                    sheet['E' + str(i + 2)] = food[names_tovars[i]][2]
            file_excel.save('food2.xlsx')     # сохраняю файл
            file = open('food2.xlsx', 'rb')
            bot.send_document(message.chat.id, file)  # отправляю файл
        else:
            bot.send_message(message.chat.id, 'Список продуктов пуст')

    elif message.text == 'Сравнить цены':    # кнопка для отправки пользователя в таблицу цен подкатегорий продукта
        bot.send_message(message.chat.id, 'Ввведите "#{номер продукта}"')

    elif message.text == '✍ EDIT':
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)  # кнопки
        item1 = types.KeyboardButton('Добавить/Уменьшить продукт')
        item2 = types.KeyboardButton('Удалить продукт')
        item3 = types.KeyboardButton('Обновить цены')
        item4 = types.KeyboardButton('Назад ⬅')
        markup.add(item1, item2, item3, item4)
        bot.send_message(message.chat.id, 'Вы зашли в редактор ✏', reply_markup=markup)


    # условие для кнопки 'Сравнить цены'
    elif message.text[0] == '#' and message.text[1:].isdigit() and int(message.text[1:]) <= len(names_tovars):
        index_ssilki = int(message.text[1]) - 1
        bot.send_message(message.chat.id, 'http://procen.ru/ceny/' + ssilki[index_ssilki])
        webbrowser.open('http://procen.ru/ceny/' + str(ssilki[index_ssilki]))    # открываю автоматически пользователю ссылку (на телефоне не работает автоматическое открытие ссылки)

    elif message.text == 'Добавить/Уменьшить продукт':  # изменить, добавить кол-во продукта, сам продукт
        bot.send_message(message.chat.id, 'Чтобы добавить продукт в список\n'
                                        'введите, например:\n'
                                        '"Абрикосы + 1 шт"\n')
        bot.send_message(message.chat.id, 'Чтобы уменьшить продукт\n'
                                        'введите, например:\n'
                                        '"Абрикосы + -1 шт"\n')

    elif message.text == 'Удалить продукт':     # удалить продукт полностью из списка
        bot.send_message(message.chat.id, 'Для удаления продукта из списка введите номер продукта')


    elif message.text == 'Назад ⬅':
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton('📋 Список продуктов')
        item2 = types.KeyboardButton('Сравнить цены')
        item3 = types.KeyboardButton('Необходимо купить')
        item4 = types.KeyboardButton('✍ EDIT')
        markup.add(item1, item2, item3, item4)
        bot.send_message(message.chat.id, 'Вы вернулись назад ⬅', reply_markup=markup)


    # Условие для удаления продукта
    elif message.text.isdigit() and int(message.text) <= len(names_tovars):
        tovar_name = names_tovars[int(message.text) - 1]
        del names_tovars[int(message.text) - 1]
        del food[tovar_name]
        bot.send_message(message.chat.id, 'Продукт был удален!')

    # кнопка для вывода списка продуктов с нулевым кол-вом
    elif message.text == 'Необходимо купить':
        file_excel = openpyxl.load_workbook(filename='food1.xlsx', data_only=True) # открываем файл ексель
        sheet = file_excel['Sheet1']
        sheet['A1'], sheet['B1'], sheet['C1'], sheet['D1'], sheet['E1'] = 'Номер', 'Название продукта', 'Кол-во', 'Ед. Измерения', 'Цена'
        count = 0
        for tovar_name in food:
            if food[tovar_name][0] == 0:
                count += 1
                sheet['A' + str(count + 1)] = count
                sheet['B' + str(count + 1)] = tovar_name
                sheet['C' + str(count + 1)] = food[tovar_name][0]
                if len(food[tovar_name]) == 3:
                    sheet['D' + str(count + 1)] = food[tovar_name][2].split('/')[1]
                    sheet['E' + str(count + 1)] = food[tovar_name][2]
                else:
                    sheet['D' + str(count + 1)] = food[tovar_name][1]
        file_excel.save('food_need.xlsx')   # сохраняем файл
        file = open('food_need.xlsx', 'rb')
        bot.send_document(message.chat.id, file)  # отправляем файл

    # Условие для обновления цен с сайта
    elif message.text == 'Обновить цены':
        bot.send_message(message.chat.id, 'Подождите примерно 2 мин....')
        for i in range(len(names_tovars)):
            if names_tovars[i] in prices:
                prod = requests.get('http://procen.ru/ceny/' + prices[names_tovars[i]] + '/')
                prod1 = BS(prod.content, 'html.parser')
                for cena in prod1.select('.price_right')[0:1]:  # парсим цену товара
                    for ed_izmeren in prod1.select('.mob_off')[1:2]:  # ед. измерения кол-ва продукта
                        if len(food[names_tovars[i]]) == 2:
                            food[names_tovars[i]].append(ed_izmeren.text + 'p/' + ed_izmeren.text)
                        else:
                            food[names_tovars[i]][2] = cena.text + 'p/' + ed_izmeren.text
                        if len(food) > 0:
                            file_excel = openpyxl.load_workbook(filename='food1.xlsx', data_only=True)
                            sheet = file_excel['Sheet1']
                            sheet['A1'], sheet['B1'], sheet['C1'], sheet['D1'], sheet['E1'] = 'Номер', 'Название продукта', 'Кол-во', 'Ед. Измерения', 'Цена'
                            for i in range(len(food)):
                                sheet['A' + str(i + 2)] = i + 1
                                yellowFill = PatternFill(start_color='F9DA5C', end_color='F9DA5C', fill_type='solid')  # желтый цвет
                                if int(food[names_tovars[i]][0]) == 0:
                                    work_sheet = sheet['A' + str(i + 2)]
                                    work_sheet.fill = yellowFill   # выделяем ячейку, если кол-во продукта равно 0
                                sheet['B' + str(i + 2)] = names_tovars[i]
                                sheet['C' + str(i + 2)] = int(food[names_tovars[i]][0])
                                if len(food[names_tovars[i]]) == 2:
                                    sheet['D' + str(i + 2)] = food[names_tovars[i]][1]
                                elif len(food[names_tovars[i]]) == 3:
                                    price_ed_izmeren = food[names_tovars[i]][2].split('/')
                                    sheet['D' + str(i + 2)] = price_ed_izmeren[1]
                                    sheet['E' + str(i + 2)] = food[names_tovars[i]][2]
                        else:
                            bot.send_message(message.chat.id, 'Список продуктов пуст')
        file_excel.save('food2.xlsx')   # сохраняем ексель
        f = open('food2.xlsx', 'rb')
        bot.send_document(message.chat.id, f)    # отправляем ексель


    # Условие для Добавления/Изменения кол-ва и продукта
    elif len(message.text.split()) == 4:
        m = message.text.split()
        if m[0] not in names_tovars:
            bot.send_message(message.chat.id, 'Продукт был добавлен в список!')
            names_tovars.append(m[0])
            food[m[0]] = [int(m[2]), m[3]]
            if food[m[0]][0] < 0:
                food[m[0]][0] = 0
        else:
            if m[2][0] == '-':

                bot.send_message(message.chat.id, 'Продукт был уменьшен!')
            else:
                bot.send_message(message.chat.id, 'Продукт был увеличен!')
            food[m[0]][0] += int(m[2])
            food[m[0]][1] = m[3]
            if food[m[0]][0] < 0:
                food[m[0][0]] = 0
    else:
        bot.send_message(message.chat.id, 'Некорректная команда!')

# планируется составить список блюд с указанием ингредиентов.
# При выборе блюда, в диалоге с ботом, указывается возможность приготовления блюда и достаточно
# ли кол-во ингредиентов. После приготовления блюда будет автоматически внесены изменения в список
# продуктов. Автоматизация изменений в списке продуктов после покупок пока в разработке.


bot.polling(none_stop=True, interval=0)

